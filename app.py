import streamlit as st
import datetime
import os
import math
import cv2
import face_recognition
import numpy as np
import requests
import pandas as pd
from supabase import create_client, Client
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────
#  CONFIG
# ─────────────────────────────────────────
SUPABASE_URL = "https://kbhajtlivcztfflmkeia.supabase.co"
SUPABASE_KEY = "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6ImtiaGFqdGxpdmN6dGZmbG1rZWlhIiwicm9sZSI6ImFub24iLCJpYXQiOjE3NzY4Njk4MDcsImV4cCI6MjA5MjQ0NTgwN30.37c8PjQYAG1ztiQJfg19RQAI3SUSF803treggp0pHp4"
EXCEL_FOLDER = "attendance_exports"
EXCEL_FILE   = os.path.join(EXCEL_FOLDER, "attendance.xlsx")
ATTENDANCE_START = datetime.time(9, 0)
ATTENDANCE_ENDS  = datetime.time(19, 0)
LATE_TIME        = datetime.time(10, 30)
ABSENT_TIME      = datetime.time(19, 0)

# India public holidays — add more as needed
INDIA_HOLIDAYS = {
    datetime.date(2025, 1, 26), datetime.date(2025, 3, 17),
    datetime.date(2025, 4, 14), datetime.date(2025, 4, 18),
    datetime.date(2025, 8, 15), datetime.date(2025, 10, 2),
    datetime.date(2025, 10, 24), datetime.date(2025, 11, 5),
    datetime.date(2025, 12, 25),
    datetime.date(2026, 1, 26), datetime.date(2026, 8, 15),
    datetime.date(2026, 10, 2), datetime.date(2026, 12, 25),
}

# ─────────────────────────────────────────
#  SUPABASE
# ─────────────────────────────────────────
@st.cache_resource
def init_supabase() -> Client:
    return create_client(SUPABASE_URL, SUPABASE_KEY)

supabase = init_supabase()

# ─────────────────────────────────────────
#  HELPERS
# ─────────────────────────────────────────
def working_days_in_range(start: datetime.date, end: datetime.date) -> int:
    """Mon–Sat days, excluding public holidays."""
    count, d = 0, start
    while d <= end:
        if d.weekday() != 6 and d not in INDIA_HOLIDAYS:
            count += 1
        d += datetime.timedelta(days=1)
    return count

def attendance_pct(present: int, working: int) -> float:
    return round((present / working) * 100, 1) if working else 0.0

def arc_path(cx, cy, r, start_deg, end_deg):
    s  = math.radians(start_deg - 90)
    e  = math.radians(end_deg   - 90)
    x1, y1 = cx + r * math.cos(s), cy + r * math.sin(s)
    x2, y2 = cx + r * math.cos(e), cy + r * math.sin(e)
    large  = 1 if (end_deg - start_deg) > 180 else 0
    return f"M {x1:.2f} {y1:.2f} A {r} {r} 0 {large} 1 {x2:.2f} {y2:.2f}"

def mark_auto_absent():
    today     = datetime.date.today()
    today_str = str(today)
    now_time  = datetime.datetime.now().time()

    # Don't run on Sundays or Holidays
    if today.weekday() == 6:
        return
    if today in INDIA_HOLIDAYS:
        return

    # Only assign auto-absent if shift attendance is over
    if now_time < ABSENT_TIME:
        return

    try:
        users = supabase.table("users").select("username", "department").execute()
        if not users.data:
            return

        for user in users.data:
            name = user.get("username")
            dept = user.get("department")

            existing = supabase.table("attendance")\
                .select("*")\
                .eq("name", name)\
                .eq("date", today_str)\
                .execute()

            # Mark absent if no login record found
            if not existing.data:
                supabase.table("attendance").insert({
                    "name":      name,
                    "date":      today_str,
                    "time":      "00:00:00",
                    "marked_by": "system",
                    "department": dept,
                    "status":    "absent"
                }).execute()
    except Exception as e:
        print("Auto absent check error:", e)

# ─────────────────────────────────────────
#  EXCEL AUTO-SAVE  (with Status column)
# ─────────────────────────────────────────
def save_attendance_excel():
    try:
        os.makedirs(EXCEL_FOLDER, exist_ok=True)
        data    = supabase.table("attendance").select("*").order("date", desc=True).order("time", desc=True).execute()
        records = data.data or []

        wb = Workbook(); ws = wb.active; ws.title = "Attendance"
        hdr_font  = Font(name="Arial", bold=True, color="FFFFFF", size=11)
        hdr_fill  = PatternFill("solid", start_color="1D3557")
        ca        = Alignment(horizontal="center", vertical="center")
        la        = Alignment(horizontal="left",   vertical="center")
        bdr       = Border(
            left=Side(style="thin", color="CCCCCC"), right=Side(style="thin", color="CCCCCC"),
            top=Side(style="thin",  color="CCCCCC"), bottom=Side(style="thin", color="CCCCCC")
        )
        alt_fill  = PatternFill("solid", start_color="F0F4F8")
        body_font = Font(name="Arial", size=10)

        ws.merge_cells("A1:G1"); ws["A1"] = "Attendance Record"
        ws["A1"].font = Font(name="Arial", bold=True, size=14, color="1D3557"); ws["A1"].alignment = ca
        ws.row_dimensions[1].height = 30

        ws.merge_cells("A2:G2")
        ws["A2"] = f"Generated: {datetime.datetime.now().strftime('%d %b %Y, %I:%M %p')}"
        ws["A2"].font = Font(name="Arial", size=9, color="888888"); ws["A2"].alignment = ca
        ws.row_dimensions[2].height = 16; ws.append([])

        headers = ["#", "Name", "Date", "Time", "Marked By", "Department", "Status"]
        ws.append(headers)
        for ci in range(1, 8):
            c = ws.cell(row=4, column=ci)
            c.font = hdr_font; c.fill = hdr_fill; c.alignment = ca; c.border = bdr
        ws.row_dimensions[4].height = 22

        for i, r in enumerate(records, 1):
            st_val = str(r.get("status", "present")).title()
            if r.get("marked_by") == "system":
                st_val = "Absent"

            ws.append([i, r.get("name", ""), r.get("date", ""),
                       str(r.get("time", ""))[:8], r.get("marked_by", ""),
                       r.get("department", ""), st_val])
            rn = 4 + i
            for ci in range(1, 8):
                c = ws.cell(row=rn, column=ci)
                c.font = body_font; c.border = bdr
                c.alignment = la if ci == 2 else ca
                if i % 2 == 0: c.fill = alt_fill
            ws.row_dimensions[rn].height = 18

        sr = 4 + len(records) + 1
        ws.cell(row=sr, column=1).value = "Total"
        ws.cell(row=sr, column=1).font  = Font(name="Arial", bold=True, size=10, color="1D3557")
        ws.cell(row=sr, column=2).value = f"=COUNTA(B5:B{4+len(records)})"
        ws.cell(row=sr, column=2).font  = Font(name="Arial", bold=True, size=10, color="1D3557")
        ws.cell(row=sr, column=2).alignment = ca

        for ci, w in enumerate([5, 22, 14, 12, 16, 18, 12], 1):
            ws.column_dimensions[get_column_letter(ci)].width = w
        ws.freeze_panes = "A5"
        wb.save(EXCEL_FILE)
        return True, EXCEL_FILE
    except Exception as e:
        return False, str(e)

# ─────────────────────────────────────────
#  SESSION & AUTO ABSENT SYSTEM
# ─────────────────────────────────────────
for k, v in [("logged_in", False), ("username", ""), ("edit_mode", False)]:
    if k not in st.session_state: st.session_state[k] = v

# Execute only when page reloads normally to mark absentees
mark_auto_absent()

# ─────────────────────────────────────────
#  MENU
# ─────────────────────────────────────────
if st.session_state.logged_in:
    menu = st.sidebar.radio("Menu", [
        "👤 Personal Dashboard", "📸 Take Photo",
        "📌 Mark Attendance",    "📊 Database", "🚪 Logout"
    ])
else:
    menu = st.sidebar.selectbox("Menu", ["Login", "Signup"])

st.title("REAL-TIME FACE ATTENDANCE SYSTEM")

# ─────────────────────────────────────────
#  SIGNUP
# ─────────────────────────────────────────
if menu == "Signup":
    st.subheader("Create Account")
    user = st.text_input("Username").strip().lower()
    pwd  = st.text_input("Password", type="password")
    if st.button("Signup"):
        if user and pwd:
            ex = supabase.table("users").select("username").eq("username", user).execute()
            if ex.data: st.warning("Username already exists.")
            else:
                supabase.table("users").insert({"username": user, "password": pwd}).execute()
                st.success("Account created!")
        else: st.warning("Fill in all fields.")

# ─────────────────────────────────────────
#  LOGIN
# ─────────────────────────────────────────
elif menu == "Login":
    st.subheader("Login")
    user = st.text_input("Username").strip().lower()
    pwd  = st.text_input("Password", type="password")
    if st.button("Login"):
        if user and pwd:
            res = supabase.table("users").select("*").eq("username", user).execute()
            if res.data and res.data[0]["password"] == pwd:
                st.session_state.logged_in = True
                st.session_state.username  = user
                st.rerun()
            else: st.error("Invalid credentials.")
        else: st.warning("Enter both fields.")

# ─────────────────────────────────────────
#  PERSONAL DASHBOARD
# ─────────────────────────────────────────
elif menu == "👤 Personal Dashboard":
    username  = st.session_state.username
    u_resp    = supabase.table("users").select("*").eq("username", username).execute()
    user_info = u_resp.data[0] if u_resp.data else {}
    emp_id    = user_info.get("emp_id", "") or ""
    dept      = user_info.get("department", "") or ""

    st.subheader("👤 My Profile")

    if "profile_success" in st.session_state:
        st.success(st.session_state.profile_success)
        del st.session_state["profile_success"]

    col_a, col_b = st.columns([3, 1])
    with col_a:
        st.markdown(f"**Username:** {username}")
        st.markdown(f"**Employee ID:** {emp_id if emp_id else '—'}")
        st.markdown(f"**Department:** {dept if dept else '—'}")
    with col_b:
        if st.button("✏️ Edit Profile"):
            st.session_state.edit_mode = not st.session_state.edit_mode

    if st.session_state.edit_mode:
        with st.form("edit_profile"):
            st.markdown("##### Update Details")
            new_emp  = st.text_input("Employee ID",  value=emp_id,  placeholder="e.g. EMP-001")
            new_dept = st.text_input("Department",   value=dept,    placeholder="e.g. Engineering")
            new_pwd  = st.text_input("New Password (leave blank to keep current)", type="password")
            s_col, c_col = st.columns(2)
            with s_col: save   = st.form_submit_button("💾 Save Changes")
            with c_col: cancel = st.form_submit_button("Cancel")

            if save:
                upd = {"emp_id": new_emp, "department": new_dept}
                if new_pwd: upd["password"] = new_pwd
                try:
                    res = supabase.table("users").update(upd).eq("username", username).execute()
                    if not res.data:
                        st.error("Update failed! Check RLS policy allows UPDATE operations for anon role.")
                    else:
                        st.session_state.profile_success = "Profile updated successfully!"
                        st.session_state.edit_mode = False
                        st.rerun()
                except Exception as e:
                    st.error(f"Database error: {e}")

            if cancel:
                st.session_state.edit_mode = False
                st.rerun()

    st.divider()

    att_resp = supabase.table("attendance").select("*").eq("name", username).execute()
    if not att_resp.data:
        st.info("No attendance records found yet.")
        st.stop()

    df = pd.DataFrame(att_resp.data)
    df["date"] = pd.to_datetime(df["date"]).dt.date

    if "status" not in df.columns:
        df["status"] = "present"
    else:
        df["status"] = df["status"].fillna("present")
    df.loc[df["marked_by"] == "system", "status"] = "absent"

    # If both an absent (system) and present row exist for the same day,
    # keep the present one. Sort: present=0 (first), absent=1 (dropped).
    df["_priority"] = df["status"].apply(lambda s: 0 if s == "present" else 1)
    df = df.sort_values("_priority").drop_duplicates(subset="date", keep="first").drop(columns="_priority")
    df["year"] = pd.to_datetime(df["date"]).dt.year

    today = datetime.date.today()
    years = sorted(df["year"].unique(), reverse=True)

    st.subheader("📊 Attendance Statistics")
    selected_year = st.selectbox("Select Year", years, index=0)

    df_year       = df[df["year"] == selected_year]
    year_start    = datetime.date(selected_year, 1, 1)
    year_end      = min(datetime.date(selected_year, 12, 31), today)
    total_working = working_days_in_range(year_start, year_end)

    present_df   = df_year[df_year["status"] != "absent"]
    present_days = len(present_df)
    absent_days  = max(total_working - present_days, 0)

    holidays_cnt = sum(1 for d in INDIA_HOLIDAYS if d.year == selected_year and year_start <= d <= year_end)
    pct          = attendance_pct(present_days, total_working)

    m1, m2, m3, m4 = st.columns(4)
    m1.metric("✅ Present Days",  present_days)
    m2.metric("❌ Absent Days",   absent_days)
    m3.metric("🎌 Holidays",      holidays_cnt)
    m4.metric("📈 Attendance %",  f"{pct}%",
              delta="Good ↑" if pct >= 75 else "Low ↓",
              delta_color="normal" if pct >= 75 else "inverse")

    st.divider()
    st.subheader("📈 Visual Breakdown")

    chart_col1, chart_col2 = st.columns(2)

    with chart_col1:
        st.markdown("**Attendance Overview**")
        total_d      = present_days + absent_days
        present_frac = (present_days / total_d) if total_d > 0 else 0
        p_end        = present_frac * 360

        cx, cy, ro, ri = 110, 110, 85, 54

        if present_days > 0 and absent_days > 0:
            slices_svg = (
                f"<path d='{arc_path(cx,cy,ro,0,p_end)} "
                f"L {cx+ri*math.cos(math.radians(p_end-90)):.2f} {cy+ri*math.sin(math.radians(p_end-90)):.2f} "
                f"{arc_path(cx,cy,ri,p_end,0)[2:]} Z' fill='#1D9E75'/>"
                f"<path d='{arc_path(cx,cy,ro,p_end,360)} "
                f"L {cx+ri*math.cos(math.radians(360-90)):.2f} {cy+ri*math.sin(math.radians(360-90)):.2f} "
                f"{arc_path(cx,cy,ri,360,p_end)[2:]} Z' fill='#E24B4A'/>"
            )
        elif present_days > 0:
            slices_svg = f"<circle cx='{cx}' cy='{cy}' r='{ro}' fill='#1D9E75'/>"
        else:
            slices_svg = f"<circle cx='{cx}' cy='{cy}' r='{ro}' fill='#E24B4A'/>"

        donut_svg = f"""
<svg viewBox="0 0 220 210" xmlns="http://www.w3.org/2000/svg" width="220" height="210">
  {slices_svg}
  <circle cx="{cx}" cy="{cy}" r="{ri}" fill="white"/>
  <text x="{cx}" y="{cy-10}" text-anchor="middle" font-size="20" font-weight="bold" fill="#1D3557">{pct}%</text>
  <text x="{cx}" y="{cy+12}" text-anchor="middle" font-size="10" fill="#888888">Attendance</text>
  <rect x="18"  y="198" width="11" height="11" fill="#1D9E75" rx="2"/>
  <text x="34"  y="207" font-size="10" fill="#444444">Present ({present_days}d)</text>
  <rect x="120" y="200" width="11" height="11" fill="#E24B4A" rx="2"/>
  <text x="136" y="207" font-size="10" fill="#444444">Absent ({absent_days}d)</text>
</svg>"""
        st.markdown(donut_svg, unsafe_allow_html=True)

    with chart_col2:
        st.markdown("**Monthly Distribution (Present Days)**")
        df_m = present_df.copy()
        if not df_m.empty:
            df_m["month"] = pd.to_datetime(df_m["date"]).dt.month
            monthly = df_m.groupby("month").size().reindex(range(1, 13), fill_value=0)
        else:
            monthly = pd.Series([0]*12, index=range(1, 13))

        mnames  = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]
        active  = [(mnames[i], int(v)) for i, v in enumerate(monthly) if v > 0]
        total_m = sum(v for _, v in active)

        pie_colors = ["#1D9E75","#378ADD","#EF9F27","#E24B4A",
                      "#7F77DD","#D85A30","#D4537E","#639922",
                      "#185FA5","#0F6E56","#BA7517","#993C1D"]

        pcx, pcy, pr = 110, 100, 78

        if total_m > 0:
            pie_paths = ""
            leg_items = ""
            angle     = 0
            for idx, (mname, val) in enumerate(active):
                sweep    = (val / total_m) * 360
                end_ang  = angle + sweep
                col      = pie_colors[idx % len(pie_colors)]
                pie_paths += (
                    f"<path d='{arc_path(pcx,pcy,pr,angle,end_ang)} "
                    f"L {pcx} {pcy} Z' fill='{col}' stroke='white' stroke-width='1.5'/>"
                )
                lx = 10 + (idx % 2) * 105
                ly = 195 + (idx // 2) * 14
                leg_items += (
                    f"<rect x='{lx}' y='{ly-9}' width='9' height='9' fill='{col}' rx='2'/>"
                    f"<text x='{lx+12}' y='{ly}' font-size='9' fill='#444444'>{mname}: {val}d</text>"
                )
                angle = end_ang

            pie_svg = f"""
<svg viewBox="0 0 220 {195 + ((len(active)-1)//2 + 1)*14 + 10}" xmlns="http://www.w3.org/2000/svg" width="220">
  {pie_paths}
  {leg_items}
</svg>"""
            st.markdown(pie_svg, unsafe_allow_html=True)
        else:
            st.info("No data for selected year.")

    st.divider()

    st.subheader("📅 Year-wise Summary (excl. Sundays & Holidays)")
    yearly = []
    for yr in years:
        df_yr = df[df["year"] == yr]
        ys    = datetime.date(yr, 1, 1)
        ye    = min(datetime.date(yr, 12, 31), today)
        wd    = working_days_in_range(ys, ye)
        pr_df = df_yr[df_yr["status"] != "absent"]
        pr    = len(pr_df)
        yearly.append({"Year": str(yr), "Present": pr, "Working Days": wd, "Attendance %": attendance_pct(pr, wd)})

    stats_df = pd.DataFrame(yearly).set_index("Year")
    st.bar_chart(stats_df[["Present", "Working Days"]])
    st.dataframe(stats_df.style.format({"Attendance %": "{:.1f}%"}), use_container_width=True)

    st.divider()

    st.subheader("📋 Records")
    show = df_year[["name", "date", "time", "status"]].sort_values("date", ascending=False)
    show["status"] = show["status"].str.upper()
    st.dataframe(show, use_container_width=True, hide_index=True)

# ─────────────────────────────────────────
#  TAKE PHOTO
# ─────────────────────────────────────────
elif menu == "📸 Take Photo":
    st.subheader("📸 Capture Face")
    img = st.camera_input("Take Photo")
    if img is not None:
        fn = f"{st.session_state.username}/{st.session_state.username}_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.jpg"
        try:
            supabase.storage.from_("faces").upload(fn, img.getvalue(), {"content-type": "image/jpeg"})
            st.success("✅ Photo uploaded!")
        except Exception as e:
            st.error(f"Upload failed: {e}")

# ─────────────────────────────────────────
#  MARK ATTENDANCE
# ─────────────────────────────────────────
elif menu == "📌 Mark Attendance":
    st.subheader("📌 Mark Attendance")
    try:
        files = supabase.storage.from_("faces").list(st.session_state.username)
    except Exception as e:
        st.error(f"Storage error: {e}"); st.stop()

    if not files:
        st.warning("No face images found. Upload photos first."); st.stop()

    known_encodings, known_names = [], []
    with st.spinner("Loading face dataset..."):
        for file in files:
            fname = file.get("name")
            if not fname: continue
            url = supabase.storage.from_("faces").get_public_url(f"{st.session_state.username}/{fname}")
            try:
                r   = requests.get(url, timeout=10); r.raise_for_status()
                arr = np.asarray(bytearray(r.content), dtype=np.uint8)
                img = cv2.imdecode(arr, cv2.IMREAD_COLOR)
                if img is None: continue
                enc = face_recognition.face_encodings(cv2.cvtColor(img, cv2.COLOR_BGR2RGB))
                if enc:
                    known_encodings.append(enc[0])
                    # Always use the actual logged-in username — files are from their folder only
                    known_names.append(st.session_state.username)
            except Exception: continue

    if not known_encodings:
        st.error("No valid face encodings found."); st.stop()

    st.success(f"✅ Loaded {len(known_encodings)} face(s).")
    u_resp    = supabase.table("users").select("department").eq("username", st.session_state.username).execute()
    user_dept = (u_resp.data[0].get("department", "") if u_resp.data else "") or ""

    now_time   = datetime.datetime.now().time()
    in_window  = ATTENDANCE_START <= now_time <= ATTENDANCE_ENDS
    now        = datetime.datetime.now()
    today_date = str(now.date())

    if not in_window:
        st.warning(
            f"⏰ Attendance camera is only available between "
            f"{ATTENDANCE_START.strftime('%I:%M %p')} and "
            f"{ATTENDANCE_ENDS.strftime('%I:%M %p')}. "
            f"Current time: {now_time.strftime('%I:%M %p')}"
        )
        st.stop()
    else:
        st.info("📷 Your **browser camera** will be used — works for every device connecting via ngrok.")

        # st.camera_input uses the CLIENT's browser camera (WebRTC), not the server's webcam.
        # This means each visitor's own device camera is used — correct for remote access.
        snap = st.camera_input("Point your face at the camera and click 📸 to mark attendance")

        if snap is not None:
            arr   = np.asarray(bytearray(snap.getvalue()), dtype=np.uint8)
            frame = cv2.imdecode(arr, cv2.IMREAD_COLOR)

            if frame is None:
                st.error("Could not decode the captured image. Please try again.")
                st.stop()

            # Resize for faster processing
            small = cv2.resize(frame, (0, 0), fx=0.5, fy=0.5)
            rgb_s = cv2.cvtColor(small, cv2.COLOR_BGR2RGB)
            locs  = face_recognition.face_locations(rgb_s)
            encs  = face_recognition.face_encodings(rgb_s, locs)

            if not locs:
                st.warning("⚠️ No face detected in the photo. Please try again with better lighting.")
            else:
                marked = []
                for encode, loc in zip(encs, locs):
                    matches = face_recognition.compare_faces(known_encodings, encode, tolerance=0.5)
                    dists   = face_recognition.face_distance(known_encodings, encode)
                    name    = "UNKNOWN"
                    color   = (0, 0, 255)

                    if True in matches:
                        idx   = int(np.argmin(dists))
                        name  = known_names[idx]
                        color = (0, 255, 0)

                        # Check for any existing record today
                        existing = supabase.table("attendance")\
                            .select("*")\
                            .eq("name", name)\
                            .eq("date", today_date)\
                            .execute()

                        existing_rows   = existing.data or []
                        already_present = any(
                            r.get("status") == "present" and r.get("marked_by") != "system"
                            for r in existing_rows
                        )
                        system_absent   = any(
                            r.get("marked_by") == "system" and r.get("status") == "absent"
                            for r in existing_rows
                        )

                        if name not in marked and not already_present:
                            now_dt = datetime.datetime.now()
                            db_ok  = False
                            try:
                                # Always INSERT a fresh present row.
                                # (Avoids needing UPDATE permission on Supabase anon role.)
                                # The dashboard deduplication keeps present over absent.
                                res = supabase.table("attendance").insert({
                                    "name":       name,
                                    "date":       str(now_dt.date()),
                                    "time":       str(now_dt.time()),
                                    "marked_by":  st.session_state.username,
                                    "department": user_dept,
                                    "status":     "present"
                                }).execute()
                                if res.data:
                                    db_ok = True
                                else:
                                    st.error(
                                        f"❌ DB INSERT returned no rows for **{name}**. "
                                        "Check Supabase RLS — the `anon` role may not have INSERT permission on `attendance`."
                                    )
                            except Exception as e:
                                st.error(f"❌ Database error for **{name}**: {e}")

                            # Only count as marked if DB write actually succeeded
                            if db_ok:
                                marked.append(name)
                                ok, path = save_attendance_excel()
                                st.toast(f"✅ {name} marked present — Excel {'saved' if ok else 'failed'}")

                        elif already_present:
                            st.info(f"ℹ️ {name} already marked present for today.")

                    # Draw box on the display frame (scale coords back up)
                    y1, x2, y2, x1 = [v * 2 for v in loc]
                    cv2.rectangle(frame, (x1, y1), (x2, y2), color, 2)
                    cv2.rectangle(frame, (x1, y2 - 30), (x2, y2), color, cv2.FILLED)
                    cv2.putText(frame, name, (x1 + 6, y2 - 6),
                                cv2.FONT_HERSHEY_SIMPLEX, 0.7, (255, 255, 255), 2)

                st.image(cv2.cvtColor(frame, cv2.COLOR_BGR2RGB),
                         channels="RGB", use_container_width=True,
                         caption="Recognition result")

                if marked:
                    st.success(f"✅ Marked present: {', '.join(marked)}")
                else:
                    # Face was detected but didn't match any known encoding
                    all_unknown = all(
                        not any(face_recognition.compare_faces(known_encodings, e, tolerance=0.5))
                        for e in encs
                    )
                    if all_unknown:
                        st.error("❌ Face not recognised. Make sure you've uploaded your photo first.")

# ─────────────────────────────────────────
#  DATABASE
# ─────────────────────────────────────────
elif menu == "📊 Database":
    st.subheader("📊 Attendance Records")
    c1, c2 = st.columns(2)
    with c1: fd = st.date_input("Filter by date", value=None)
    with c2: fn = st.text_input("Filter by name", placeholder="All")
    try:
        q = supabase.table("attendance").select("*").order("date", desc=True).order("time", desc=True)
        if fd: q = q.eq("date", str(fd))
        if fn: q = q.ilike("name", f"%{fn}%")
        data = q.execute()

        if data.data:
            df_db = pd.DataFrame(data.data)

            if "status" not in df_db.columns:
                df_db["status"] = "present"
            else:
                df_db["status"] = df_db["status"].fillna("present")
            df_db.loc[df_db["marked_by"] == "system", "status"] = "absent"
            df_db["status"] = df_db["status"].str.upper()

            cols  = ["id", "name", "date", "time", "status", "marked_by", "department"]
            df_db = df_db[[c for c in cols if c in df_db.columns]]

            st.dataframe(df_db, use_container_width=True)
            st.caption(f"Total: {len(data.data)} records")

            if st.button("⬇ Save / Refresh Excel now"):
                ok, res = save_attendance_excel()
                st.success(f"Saved to `{res}`") if ok else st.error(f"Failed: {res}")
        else:
            st.info("No records found.")
    except Exception as e:
        st.error(f"Error: {e}")

# ─────────────────────────────────────────
#  LOGOUT
# ─────────────────────────────────────────
elif menu == "🚪 Logout":
    st.session_state.logged_in = False
    st.session_state.username  = ""
    st.success("Logged out!")
    st.rerun() 
